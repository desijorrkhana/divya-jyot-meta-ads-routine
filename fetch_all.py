#!/usr/bin/env python3
"""
Divya Jyot — Daily data fetcher for the 7 PM agency + sales-manager routine.

Pulls and writes to data.json for Claude Code to analyse:
  (A) Meta Ads performance (spend/clicks/CPL/CTR by campaign/adset/ad)
  (B) Timed Meta leads (created_time + phone + intent) from the V3 CRM Event sheet
  (C) The team's Google Sheet (Facebook + SVD tabs) — full history

ENV VARS:
  META_AD_ACCOUNT_ID   1078185463738176  (act_ prefix optional — auto-added)
  META_ADS_TOKEN       token with ads_read
  GOOGLE_SHEET_ID      main team sheet id
  LEADS_SHEET_ID       V3 CRM Event sheet id (has a default)
  GOOGLE_SA_B64        base64 of service-account JSON (preferred)
  GOOGLE_SA_JSON       raw OR base64 service-account JSON (auto-detected)
  TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID   for --send delivery

Hardened against the three issues seen in production:
  1. META_AD_ACCOUNT_ID missing "act_" prefix      -> auto-prepended
  2. GOOGLE_SA_JSON actually base64, not raw JSON    -> auto-detected & decoded
  3. self-signed proxy cert breaks Google SSL        -> resilient HTTP via google_auth_httplib2
Runtime reduced by running the 5 Meta API calls + Google reads concurrently.
"""
import json, os, base64, ssl, time, urllib.request, urllib.parse
from datetime import date, datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor

# ---------- CONFIG ----------
def _acct():
    a = os.environ.get("META_AD_ACCOUNT_ID", "").strip()
    if a and not a.startswith("act_"):
        a = "act_" + a            # FIX 1: auto-prepend act_
    return a
AD_ACCOUNT_ID = _acct()
TOKEN = os.environ.get("META_ADS_TOKEN", "")
API = "v25.0"
IST = timezone(timedelta(hours=5, minutes=30))
# `or` (not a .get default): CI sets unset secrets as EMPTY-STRING env vars, and an
# empty spreadsheet id must still fall back — .get's default only covers a MISSING var.
LEADS_SHEET_ID = os.environ.get("LEADS_SHEET_ID") or "1752IvdN_Qdwd36xuQp5EJ55jATaoOyv4dwzoqeIAeZY"

def day(n): return (date.today() - timedelta(days=n)).isoformat()

# tolerant HTTP GET (Meta). Short timeout so one slow call can't hang the run.
def _get(url, timeout=25):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE   # FIX 3: tolerate self-signed proxy cert in the env
    with urllib.request.urlopen(url, timeout=timeout, context=ctx) as r:
        return json.load(r)

# ---------- (A) Meta insights ----------
_INSIGHT_FIELDS = ("campaign_name,adset_name,ad_name,spend,impressions,reach,clicks,"
                   "inline_link_clicks,ctr,cpm,cpc,frequency,actions")

# Match Ads Manager's default attribution so counts line up with what you see on screen.
# Ads Manager default = 7-day click + 1-day view. We pin the same here.
# NOTE: as of API v25.0, action_attribution_windows takes flat strings (e.g. "7d_click"),
# NOT {"event_type":..., "window_days":...} objects — the old object form now 400s outright.
_ATTRIBUTION = json.dumps(["7d_click", "1d_view"])

def meta_insights(level, since, until, time_increment=None):
    if not TOKEN or not AD_ACCOUNT_ID:
        return [{"error": "Missing META_ADS_TOKEN or META_AD_ACCOUNT_ID"}]
    params = {
        "level": level, "fields": _INSIGHT_FIELDS + ",date_start",
        "time_range": json.dumps({"since": since, "until": until}),
        "action_attribution_windows": _ATTRIBUTION,
        "use_unified_attribution_setting": "true",   # use the ad set's configured attribution
        "limit": "500", "access_token": TOKEN,
    }
    if time_increment:
        params["time_increment"] = str(time_increment)   # 1 = one row per day (for trends)
    url = f"https://graph.facebook.com/{API}/{AD_ACCOUNT_ID}/insights?" + urllib.parse.urlencode(params)
    try:
        return _get(url).get("data", [])
    except Exception as e:
        return [{"error": str(e)}]

# CANONICAL lead count. Meta returns SEVERAL overlapping lead action types for instant-form
# ads (e.g. "lead", "leadgen_grouped", "onsite_conversion.lead_grouped", "onsite_lead").
# Summing all of them (the old bug) multi-counted — turning 4 real leads into ~20 and
# crushing CPL from ~₹100 to ~₹19. We pick ONE canonical type, preferring the grouped
# instant-form action that matches Ads Manager's "Leads" column, and never add the others.
_LEAD_PRIORITY = [
    "onsite_conversion.lead_grouped",   # instant-form leads (what Ads Manager shows)
    "leadgen_grouped",
    "onsite_lead",
    "lead",
]

def leads_from(actions):
    if not actions:
        return 0
    by_type = {}
    for a in actions:
        t = a.get("action_type", "")
        try:
            by_type[t] = by_type.get(t, 0) + int(float(a.get("value", 0) or 0))
        except (ValueError, TypeError):
            pass
    # Return the FIRST canonical type present — do NOT sum across types.
    for t in _LEAD_PRIORITY:
        if t in by_type:
            return by_type[t]
    # fallback: if Meta used some other single lead-ish type, take the max single one
    lead_like = {t: v for t, v in by_type.items() if "lead" in t}
    return max(lead_like.values()) if lead_like else 0

def shape_meta(rows):
    out = []
    for r in rows:
        if "error" in r:
            out.append(r); continue
        spend = float(r.get("spend", 0) or 0); leads = leads_from(r.get("actions"))
        # keep the raw lead-ish action breakdown so the report (and we) can verify the
        # count is right and not double-counting. This is the audit trail.
        raw_leadish = {a.get("action_type"): a.get("value")
                       for a in (r.get("actions") or []) if "lead" in a.get("action_type","")}
        out.append({
            "campaign": r.get("campaign_name"), "adset": r.get("adset_name"), "ad": r.get("ad_name"),
            "date": r.get("date_start"),
            "spend": round(spend,2), "leads": leads,
            "cpl": round(spend/leads,2) if leads else None,
            "lead_actions_raw": raw_leadish,   # audit: all lead-type actions Meta returned
            "impressions": int(r.get("impressions",0) or 0),
            "reach": int(r.get("reach",0) or 0),
            "clicks": int(r.get("clicks",0) or 0),
            "link_clicks": int(r.get("inline_link_clicks",0) or 0),
            "ctr": round(float(r.get("ctr",0) or 0),2),
            "cpc": round(float(r.get("cpc",0) or 0),2),
            "cpm": round(float(r.get("cpm",0) or 0),2),
            "frequency": round(float(r.get("frequency",0) or 0),2),
        })
    return out

# ---------- (B) timed leads from the V3 CRM Event sheet ----------
def normalize_phone(p):
    if not p: return ""
    d = "".join(ch for ch in str(p) if ch.isdigit())
    return d[-10:] if len(d) >= 10 else d

def parse_lead_rows(rows, source_tab=""):
    # NOTE: the "V3 CRM Event" spreadsheet (LEADS_SHEET_ID) has ONE tab per lead FORM, not
    # one tab total. Sheet1 = the Studio form, Sheet2 = the "2BHK" form added when that
    # campaign launched (2026-07-06) — each new form gets its own new tab. Missing a tab
    # here silently drops an entire campaign's leads (happened: Sheet2 was never read,
    # making the whole 2BHK campaign look untracked when it wasn't). ALWAYS enumerate every
    # tab in the spreadsheet (see fetch_sheets) rather than hardcoding tab names/count.
    if not rows or len(rows) < 2:
        return []
    header = [h.strip().lower() for h in rows[0]]
    def col(*names):
        for n in names:
            try: return header.index(n)
            except ValueError: pass
        return None
    ci = {
        "created_time": col("created_time"), "ad_name": col("ad_name"),
        "adset_name": col("adset_name"), "campaign_name": col("campaign_name"),
        "platform": col("platform"),
        # intent question wording differs per form/tab — accept either.
        "intent": col("when_are_you_planning_to_purchase?", "when_are_you_planning_to_buy"),
        "budget": col("what_is_your_budget_for_this_purchase?"),  # only present on some forms
        "full_name": col("full_name"), "phone_number": col("phone_number"),
        "lead_status": col("lead_status"),
    }
    out = []
    for r in rows[1:]:
        def g(key):
            i = ci.get(key)
            return r[i].strip() if (i is not None and i < len(r)) else ""
        ct = g("created_time"); name = g("full_name")
        if not ct or "test lead" in name.lower() or "dummy data" in name.lower():
            continue
        try:
            ct_ist = datetime.fromisoformat(ct).astimezone(IST).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            ct_ist = ct
        out.append({
            "created_time_raw": ct, "created_time_ist": ct_ist,
            "phone10": normalize_phone(g("phone_number")), "name": name,
            "platform": g("platform"), "intent": g("intent"), "budget": g("budget"),
            "ad": g("ad_name"), "adset": g("adset_name"), "campaign": g("campaign_name"),
            "lead_status": g("lead_status"), "source_tab": source_tab,
        })
    return out

# ---------- GOOGLE Sheets ----------
def _decode_sa():
    """FIX 2: accept base64 in EITHER var, or raw JSON. Auto-detect."""
    raw = os.environ.get("GOOGLE_SA_B64", "") or os.environ.get("GOOGLE_SA_JSON", "")
    if not raw:
        raise RuntimeError("Missing GOOGLE_SA_B64 / GOOGLE_SA_JSON")
    raw = raw.strip()
    if raw.startswith("{"):
        return json.loads(raw)                       # already raw JSON
    try:
        return json.loads(base64.b64decode(raw).decode("utf-8"))   # base64 -> JSON
    except Exception:
        return json.loads(raw)                       # last resort

def _google_creds():
    from google.oauth2.service_account import Credentials
    sa_info = _decode_sa()
    # drive.readonly is needed on top of spreadsheets.readonly so we can read the sheet's
    # REVISION HISTORY (Drive API) — that's the only place edit *times* live, since the
    # team writes only dates (never clock times) into the cells themselves.
    return Credentials.from_service_account_info(
        sa_info, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly",
                         "https://www.googleapis.com/auth/drive.readonly"])

def google_sheets(creds=None):
    from googleapiclient.discovery import build
    import google_auth_httplib2, httplib2
    creds = creds or _google_creds()
    # FIX 3: build an http object that tolerates the env's self-signed proxy cert.
    # NOTE: httplib2.Http is NOT thread-safe — each concurrent caller MUST get its own
    # instance, or interleaved requests corrupt the shared TLS session (intermittent
    # "decryption failed or bad record mac" errors under concurrency).
    http = google_auth_httplib2.AuthorizedHttp(creds, http=httplib2.Http(disable_ssl_certificate_validation=True))
    return build("sheets", "v4", http=http, cache_discovery=False)

def fetch_sheets():
    sheet_id = os.environ.get("GOOGLE_SHEET_ID", "")
    if not sheet_id:
        return {"error": "Missing GOOGLE_SHEET_ID"}
    try:
        creds = _google_creds()
    except Exception as e:
        return {"error": f"google auth failed: {e}"}

    # Sheets API occasionally 503s ("service is currently unavailable") — transient, seen
    # in production 2026-08-03 wiping facebook_tab + svd_tab for a whole run. Retry a few
    # times with backoff before giving up, same spirit as the Drive revision pacing below.
    def read(sid, rng, attempts=4):
        last = None
        for i in range(attempts):
            try:
                svc = google_sheets(creds)   # fresh http per thread — see NOTE above
                return svc.spreadsheets().values().get(spreadsheetId=sid, range=rng).execute().get("values", [])
            except Exception as e:
                last = e
                if i < attempts - 1:
                    time.sleep(3 * (i + 1))
        return [["error", str(last)]]

    # The CRM Event spreadsheet gets a NEW TAB per lead form (Sheet1 = Studio, Sheet2 =
    # "2BHK" added 2026-07-06, and any future form gets its own tab too) — enumerate every
    # tab rather than hardcoding "Sheet1", or a whole new campaign's leads silently vanish.
    try:
        svc0 = google_sheets(creds)
        crm_tabs = [s["properties"]["title"]
                    for s in svc0.spreadsheets().get(spreadsheetId=LEADS_SHEET_ID).execute().get("sheets", [])]
    except Exception as e:
        crm_tabs = ["Sheet1"]  # fall back to the known-good tab if listing itself fails
        crm_tabs_error = str(e)
    else:
        crm_tabs_error = None

    # Run the sheet/tab reads concurrently (independent network calls).
    with ThreadPoolExecutor(max_workers=2 + len(crm_tabs)) as ex:
        f_fb   = ex.submit(read, sheet_id, "Facebook!A1:N2000")
        f_svd  = ex.submit(read, sheet_id, "SVD!A1:O500")
        f_leads = {tab: ex.submit(read, LEADS_SHEET_ID, f"{tab}!A1:Z5000") for tab in crm_tabs}
        fb, svd = f_fb.result(), f_svd.result()
        leads_by_tab = {tab: fut.result() for tab, fut in f_leads.items()}

    leads_parsed, leads_errors = [], {}
    for tab, raw in leads_by_tab.items():
        if raw and raw[0] and raw[0][0] == "error":
            leads_errors[tab] = raw[0][1]
            continue
        leads_parsed.extend(parse_lead_rows(raw, source_tab=tab))
    err = leads_errors or crm_tabs_error
    return {
        "facebook_tab": fb,
        "svd_tab": svd,
        "meta_leads_crm_tabs": crm_tabs,
        "meta_leads_timed": leads_parsed,
        "meta_leads_error": err,
    }

# ---------- CONTACT TIMES via Drive revision history ----------
# The team writes only DATES (never clock times) in the sheet, so the cells alone can't
# say WHEN a lead was first entered/called. But every edit creates a Drive revision with a
# modifiedTime. We export a handful of recent revisions of the team sheet, parse the
# Facebook tab in each, and bracket for each fresh Meta lead: when its row first appeared,
# and when its first feedback text appeared. Resolution = gap between adjacent revisions
# (typically minutes-to-hours on an actively edited sheet) instead of a whole day.
REVISION_LOOKBACK_DAYS = int(os.environ.get("REVISION_LOOKBACK_DAYS", "2"))
_MAX_REVISION_DOWNLOADS = 24
_REVISION_STAGE_DEADLINE_S = 360   # hard cap for the whole stage — a daily job must not hang
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _drive_http(creds):
    import google_auth_httplib2, httplib2
    return google_auth_httplib2.AuthorizedHttp(
        creds, http=httplib2.Http(disable_ssl_certificate_validation=True))

def _cell_phone(c):
    # xlsx exports numeric-looking phones as floats (7506401771.0) — the ".0" corrupts
    # normalize_phone's last-10-digits logic, so collapse integral floats first.
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():
        c = int(c)
    return normalize_phone(c)

def _parse_fb_tab_xlsx(blob, phones):
    """From one exported revision, return {phone10: {'present': bool, 'feedback': bool}}."""
    import io, openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    if "Facebook" not in wb.sheetnames:
        return {}
    state = {p: {"present": False, "feedback": False} for p in phones}
    for row in wb["Facebook"].iter_rows(max_col=14, values_only=True):
        p = _cell_phone(row[5]) if len(row) > 5 else ""
        if p in state:
            state[p]["present"] = True
            # feedback = any text in Feedback (col H, idx 7) or the follow-up columns
            if any(c is not None and str(c).strip() for c in row[7:14]):
                state[p]["feedback"] = True
    wb.close()
    return state

def fetch_contact_history(leads):
    """leads = meta_leads_timed. Only leads that arrived inside the lookback window are
    tracked (older ones were in the sheet before the earliest revision we scan)."""
    sheet_id = os.environ.get("GOOGLE_SHEET_ID", "")
    if not sheet_id:
        return {"error": "Missing GOOGLE_SHEET_ID"}
    cutoff = (datetime.now(IST) - timedelta(days=REVISION_LOOKBACK_DAYS)).strftime("%Y-%m-%d %H:%M:%S")
    fresh = {l["phone10"]: l for l in leads
             if l["phone10"] and l["created_time_ist"] >= cutoff}
    if not fresh:
        return {"error": None, "window_days": REVISION_LOOKBACK_DAYS, "leads": {}}
    try:
        import openpyxl  # noqa: F401 — fail early with a clear message
    except ImportError:
        return {"error": "openpyxl not installed (pip install openpyxl) — contact times unavailable"}
    try:
        http = _drive_http(_google_creds())
        url = (f"https://www.googleapis.com/drive/v3/files/{sheet_id}/revisions"
               f"?fields=revisions(id,modifiedTime,exportLinks)&pageSize=1000")
        revs = []
        for attempt in range(3):   # seen returning 200 with an empty body transiently
            resp, content = http.request(url)
            if resp.status == 200:
                revs = json.loads(content).get("revisions", [])
                if revs:
                    break
            time.sleep(1 + attempt)
        if not revs:
            return {"error": f"revisions.list returned no revisions (last HTTP {resp.status})"}
        def ist(mt):  # RFC3339 UTC -> IST string
            return (datetime.fromisoformat(mt.replace("Z", "+00:00"))
                    .astimezone(IST).strftime("%Y-%m-%d %H:%M:%S"))
        for r in revs:
            r["ist"] = ist(r["modifiedTime"])
        revs.sort(key=lambda r: r["ist"])
        window = [r for r in revs if r["ist"] >= cutoff]
        baseline = revs[revs.index(window[0]) - 1:revs.index(window[0])] if (window and revs.index(window[0]) > 0) else []
        scan = baseline + window
        total_in_window = len(scan)
        if len(scan) > _MAX_REVISION_DOWNLOADS:
            # keep first + last, sample the middle evenly — resolution degrades gracefully
            step = (len(scan) - 1) / (_MAX_REVISION_DOWNLOADS - 1)
            scan = [scan[round(i * step)] for i in range(_MAX_REVISION_DOWNLOADS)]

        out = {p: {"name": l["name"], "meta_arrival_ist": l["created_time_ist"],
                   "row_appeared_between": None, "feedback_appeared_between": None}
               for p, l in fresh.items()}
        prev_state, prev_ist = {}, None
        scanned, failed = 0, 0
        stage_deadline = time.monotonic() + _REVISION_STAGE_DEADLINE_S
        for rv in scan:
            link = rv.get("exportLinks", {}).get(_XLSX_MIME)
            if not link:
                continue
            # Drive throttles revision exports with a small burst quota (429 after ~4 rapid
            # downloads, refills over tens of seconds) — pace + retry, but never blow the
            # stage deadline: partial coverage is reported honestly below.
            blob = None
            for attempt in range(4):
                if time.monotonic() > stage_deadline:
                    break
                resp, b = http.request(link)
                if resp.status == 200:
                    blob = b
                    break
                if resp.status != 429:
                    break
                time.sleep(min(12 + 12 * attempt, max(0, stage_deadline - time.monotonic())))
            if blob is None:
                failed += 1
                if time.monotonic() > stage_deadline:
                    break
                continue
            scanned += 1
            state = _parse_fb_tab_xlsx(blob, set(fresh))
            for p, s in state.items():
                o = out[p]
                ps = prev_state.get(p, {"present": False, "feedback": False})
                if s["present"] and not ps["present"] and o["row_appeared_between"] is None:
                    o["row_appeared_between"] = [prev_ist, rv["ist"]]
                if s["feedback"] and not ps["feedback"] and o["feedback_appeared_between"] is None:
                    o["feedback_appeared_between"] = [prev_ist, rv["ist"]]
            prev_state, prev_ist = state, rv["ist"]
            time.sleep(4)   # pacing: stay under the export burst quota
        return {"error": None, "window_days": REVISION_LOOKBACK_DAYS,
                "revisions_scanned": scanned, "revisions_failed": failed,
                "revisions_in_window": total_in_window,
                "note": ("[after, by] IST bounds from Drive revision history. null bound = state "
                         "unknown before the first scanned revision; a null field = not yet "
                         "seen in the sheet as of the newest scanned revision. A skipped "
                         "(failed) revision widens the bracket, it never fabricates one."),
                "leads": out}
    except Exception as e:
        return {"error": f"contact history failed: {e}"}

# ---------- ASSEMBLE (Meta calls + Sheets all concurrent) ----------
def run_fetch():
    jobs = {
        "today": ("campaign", day(0), day(0)),   # TODAY so far (up to run time, ~7PM)
        "today_ads": ("ad",   day(0), day(0)),    # today's per-ad, so fb vs ig split is visible
        "yc": ("campaign", day(1), day(1)),
        "ya": ("adset",    day(1), day(1)),
        "yd": ("ad",       day(1), day(1)),
        "l7": ("campaign", day(7), day(1)),
        "l30":("ad",       day(30), day(1)),
        "l30_daily": ("campaign", day(30), day(0), 1),  # per-day rows for the dashboard trend
        "l30_daily_ads": ("ad", day(30), day(0), 1),    # per-day per-ad, for ad-level range views
    }
    results = {}
    with ThreadPoolExecutor(max_workers=8) as ex:
        meta_futs = {k: ex.submit(meta_insights, *v) for k, v in jobs.items()}
        sheet_fut = ex.submit(fetch_sheets)
        for k, fut in meta_futs.items():
            results[k] = shape_meta(fut.result())
        sheet = sheet_fut.result()

    # needs the timed leads' phones, so it runs after the sheet fetch (sequential is fine —
    # it's one revisions.list + a bounded number of small xlsx exports)
    sheet["contact_history"] = fetch_contact_history(sheet.get("meta_leads_timed") or [])

    data = {
        "dates": {"yesterday": day(1), "daybefore": day(2), "today": day(0), "tz": "IST",
                  "note": "today_* covers TODAY from midnight IST up to the run time (~7PM)."},
        "meta": {
            "today_campaigns":     results["today"],
            "today_ads":           results["today_ads"],
            "yesterday_campaigns": results["yc"],
            "yesterday_adsets":    results["ya"],
            "yesterday_ads":       results["yd"],
            "last7_campaigns":     results["l7"],
            "last30_ads":          results["l30"],
            "last30_daily_campaigns": results["l30_daily"],
            "last30_daily_ads":       results["l30_daily_ads"],
        },
        "sheet": sheet,
    }
    with open(os.path.join(os.path.dirname(__file__), "data.json"), "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("Wrote data.json — Meta insights + timed leads + Facebook/SVD tabs (concurrent fetch)")

# ---------- TELEGRAM ----------
def send_telegram_file(path="report.md"):
    token = os.environ.get("TELEGRAM_BOT_TOKEN", ""); chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    if not token or not chat_id:
        print("Missing TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID"); return False
    try:
        with open(os.path.join(os.path.dirname(__file__), path), encoding="utf-8") as f:
            msg = f.read()
    except FileNotFoundError:
        print(f"No {path} to send"); return False
    chunks = [msg[i:i+3800] for i in range(0, len(msg), 3800)] or [msg]
    ok = True
    for chunk in chunks:
        # NOTE: send as PLAIN TEXT (no parse_mode). Markdown tables/symbols in the report
        # break Telegram's parser and cause silent failures — plain text is reliable.
        body = urllib.parse.urlencode({
            "chat_id": chat_id, "text": chunk, "disable_web_page_preview": "true",
        }).encode()
        try:
            req = urllib.request.Request(f"https://api.telegram.org/bot{token}/sendMessage", data=body)
            with urllib.request.urlopen(req, timeout=20) as r:
                if not json.load(r).get("ok"):
                    ok = False; print("Telegram returned not-ok")
        except Exception as e:
            ok = False; print("Telegram send failed:", e)
    print("Report sent to Telegram." if ok else "Telegram delivery had errors.")
    return ok

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--send":
        send_telegram_file("report.md")
    else:
        run_fetch()
